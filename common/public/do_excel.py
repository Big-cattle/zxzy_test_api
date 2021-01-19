from openpyxl import load_workbook
from  ZXZY_API_AUTO.common.conf.read_config import ReadConfig
from  ZXZY_API_AUTO.common.public.abs_path import *
'''读取Excel表格中的测试数据'''
class DoExcel:
    @staticmethod
    def read_excel():  #静态方法，读取表格数据
        test_data = []   #创建一个列表
        res = ReadConfig.read_cf('CON','mode')  #调用读取配置文件封装方法，读取配置文件内容
        wb = load_workbook(testdata_path)  # 定位Excel表格并赋值给wb
        for i in res: #遍历配置文件信息
            sheet = wb[i] #定位表单
            
            ''''''
            if res[i]=='all':
                for x in range(2,sheet.max_row+1):
                    excel_data = {}
                    excel_data['id'] = sheet.cell(x,1).value
                    excel_data['moudle'] = sheet.cell(x,2).value
                    excel_data['description'] = sheet.cell(x,3).value
                    excel_data['data'] = sheet.cell(x,4).value
                    excel_data['expect'] = sheet.cell(x,5).value
                    excel_data['http_method'] = sheet.cell(x,6).value
                    excel_data['url'] = sheet.cell(x,7).value
                    test_data.append(excel_data)
            else:
                for x in range(2, sheet.max_row+1):
                    excel_data = {}
                    excel_data['id'] = sheet.cell(x, 1).value
                    excel_data['moudle'] = sheet.cell(x, 2).value
                    excel_data['description'] = sheet.cell(x, 3).value
                    excel_data['data'] = sheet.cell(x, 4).value
                    excel_data['expect'] = sheet.cell(x, 5).value
                    excel_data['http_method'] = sheet.cell(x, 6).value
                    excel_data['url'] = sheet.cell(x, 7).value
                    if excel_data['id'] in res[i]:
                        test_data.append(excel_data)
        return test_data
    
    @staticmethod
    def  write_excel(sheet_n,row,result,ifpass):
        wb = load_workbook(testdata_path)
        sheet = wb[sheet_n]
        sheet.cell(row,8).value = result
        sheet.cell(row,9).value = ifpass
        wb.save(testdata_path)
    
if __name__ == '__main__':
    print(DoExcel().read_excel())